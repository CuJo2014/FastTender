"""Прогон матчера по золотому датасету (раздел 16.3 пункт 7, 20.2).

Главный артефакт Phase 1: цифра точности, по которой принимается решение
о Phase 2 (раздел 3.3, 16.6).

Запуск:
    uv run python -m fasttender.scripts.eval_gold path/to/gold.xlsx
    uv run python -m fasttender.scripts.eval_gold gold.xlsx --output result.xlsx --top-k 5

Шаблон входного файла — Приложение C.3:
    Лист «Датасет» с двухуровневой шапкой (R1 — групповые лейблы, R2 — заголовки колонок).
    Колонки находятся по заголовку, а не по позиции — допустимо переставлять
    или добавлять/убирать колонки между обязательными.

Заполняет колонки:
    - «Результат матчера: артикул»     ← top-1 catalog candidate
    - «Результат матчера: уверенность» ← confidence top-1
    - «Совпало? (да/нет)»              ← top-1 == правильный артикул каталога

Метрики (раздел 20.2) считаются по строкам со статусом «найдено» или «аналог»
и непустым «Правильный артикул каталога»:
    - Recall@K — доля строк, где правильный ответ в топ-K
    - Precision@1 — доля, где правильный ответ — на 1-м месте
    - MRR — среднее обратное место правильного ответа (0 если вне топ-K)

Дополнительно для строк со статусом «не найдено»:
    - not_found_correct — матчер тоже ничего не вернул или confidence < min_threshold
"""

from __future__ import annotations

import argparse
import asyncio
import sys
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from fasttender.core.config import get_settings
from fasttender.repositories.pg_trgm import PgTrgmSearchRepository
from fasttender.services.matcher import MatchingEngine, MatchInput, MatchResult
from fasttender.services.parser.value_normalizer import (
    clean_string,
    normalize_article,
    normalize_name,
)

# --- Конфигурация колонок (приложение C.3 + правка) ---

# Логические поля → варианты заголовков. Поиск без учёта регистра,
# по принципу «начинается с» с границей слова.
COLUMN_SYNONYMS: dict[str, tuple[str, ...]] = {
    "row_num": ("№", "номер"),
    "source_file": ("файл-источник", "файл источник"),
    "name": ("наименование (как у клиента)", "наименование клиент", "наименование"),
    "article": ("артикул (как у клиента)", "артикул клиент", "артикул"),
    "manufacturer": ("производитель (как у клиента)", "производитель клиент", "производитель"),
    "attributes": (
        "характеристика (как у клиента)",
        "характеристики (как у клиента)",
        "характеристика",
        "характеристики",
    ),
    "quantity": ("кол-во", "количество", "к-во"),
    "unit": ("ед. изм.", "ед.изм", "ед изм", "единица"),
    "expected_article": ("→ правильный артикул каталога", "правильный артикул", "артикул каталога"),
    "expected_code_1c": (
        "→ код 1с каталога",
        "→ правильный код 1с",
        "правильный код 1с",
        "код 1с каталога",
        "код каталога",
    ),
    "expected_name": ("→ правильное наименование каталога", "правильное наименование"),
    "label_status": ("статус разметки", "статус"),
    "labeler_notes": ("примечание разметчика", "примечание"),
    "result_article": ("результат матчера: артикул", "результат: артикул"),
    "result_code_1c": ("результат матчера: код 1с", "результат: код 1с"),
    "result_confidence": ("результат матчера: уверенность", "результат: уверенность"),
    "result_matched": ("совпало? (да/нет)", "совпало"),
}

# expected_article ИЛИ expected_code_1c достаточно — хотя бы один identifier
REQUIRED_COLUMNS = ("name", "label_status")
RESULT_COLUMNS = ("result_article", "result_code_1c", "result_confidence", "result_matched")

# Допустимые значения «Статус разметки»
STATUS_FOUND = {"найдено", "найден"}
STATUS_ANALOG = {"аналог"}
STATUS_NOT_FOUND = {"не найдено", "не найден"}
STATUS_UNSURE = {"сомнительно", "под вопросом", "?"}


# --- Доменные структуры ---


@dataclass
class GoldRow:
    """Одна строка золотого датасета."""

    sheet_row: int  # реальный 1-based индекс в листе
    row_num: int | None
    source_file: str | None
    name: str
    article: str | None
    manufacturer: str | None
    attributes: str | None
    quantity: str | None
    unit: str | None
    expected_article: str | None
    expected_code_1c: str | None  # альтернативный identifier когда артикула нет
    expected_name: str | None
    label_status: str
    labeler_notes: str | None

    @property
    def expected_article_normalized(self) -> str | None:
        return normalize_article(self.expected_article)

    @property
    def has_expected_identifier(self) -> bool:
        """Хотя бы один identifier (артикул или код 1С) задан."""
        return bool(self.expected_article_normalized) or bool(self.expected_code_1c)


@dataclass
class RowOutcome:
    """Результат прогона по одной строке."""

    row: GoldRow
    match_result: MatchResult
    top_articles_normalized: list[str]  # top-K catalog articles, normalized
    top_codes_1c: list[str]  # top-K catalog code_1c (raw, без нормализации)
    expected_rank: int | None  # 1-based позиция expected в top-K по любому из identifier'ов
    matched: bool  # точное совпадение top-1 по любому из identifier'ов


@dataclass
class GoldMetrics:
    """Сводные метрики по золотому датасету."""

    top_k: int

    total_rows: int = 0
    skipped_no_name: int = 0
    skipped_unsure: int = 0

    # Строки со статусом «найдено»/«аналог» + expected_article
    applicable: int = 0
    recall_at_k_hits: int = 0
    precision_at_1_hits: int = 0
    reciprocal_ranks_sum: float = 0.0

    # Строки со статусом «не найдено»
    not_found_rows: int = 0
    not_found_correct: int = 0  # матчер тоже ничего не дал

    # Распределение по статусам
    status_counts: dict[str, int] = field(default_factory=dict)

    @property
    def recall_at_k(self) -> float:
        return self.recall_at_k_hits / self.applicable if self.applicable else 0.0

    @property
    def precision_at_1(self) -> float:
        return self.precision_at_1_hits / self.applicable if self.applicable else 0.0

    @property
    def mrr(self) -> float:
        return self.reciprocal_ranks_sum / self.applicable if self.applicable else 0.0

    @property
    def not_found_precision(self) -> float:
        return self.not_found_correct / self.not_found_rows if self.not_found_rows else 0.0


# --- Парсинг шапки ---


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def detect_columns(ws: Worksheet, *, max_scan_rows: int = 5) -> tuple[int, dict[str, int]]:
    """Определяет 1-based индекс строки заголовков и mapping {field → 1-based column}.

    Сканирует первые max_scan_rows строк, выбирает ту, где найдено больше
    всего распознанных колонок (и при этом обязательно есть `name`).
    """
    best_row = 0
    best_mapping: dict[str, int] = {}
    best_score = 0

    for row_idx in range(1, min(max_scan_rows, ws.max_row) + 1):
        mapping: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            header_text = _norm_header(ws.cell(row=row_idx, column=col_idx).value)
            if not header_text:
                continue
            field_name = _match_synonym(header_text)
            if field_name and field_name not in mapping:
                mapping[field_name] = col_idx
        if "name" in mapping and len(mapping) > best_score:
            best_score = len(mapping)
            best_row = row_idx
            best_mapping = mapping

    if best_row == 0:
        raise GoldEvalError(
            "Не нашёл строку заголовков. Проверь, что лист «Датасет» содержит "
            "колонку с наименованием по образцу Приложения C.3."
        )

    missing = [c for c in REQUIRED_COLUMNS if c not in best_mapping]
    if missing:
        raise GoldEvalError(
            f"В шапке не хватает обязательных колонок: {missing}. "
            f"Распознаны: {sorted(best_mapping)}"
        )

    return best_row, best_mapping


def _match_synonym(header_text: str) -> str | None:
    for field_name, synonyms in COLUMN_SYNONYMS.items():
        for syn in synonyms:
            syn_norm = _norm_header(syn)
            if not syn_norm:
                continue
            if header_text == syn_norm:
                return field_name
            # «начинается с» с границей слова
            if header_text.startswith(syn_norm):
                tail = header_text[len(syn_norm) :]
                if not tail or tail[0] in " :-.,/(":
                    return field_name
    return None


# --- Чтение строк ---


def read_rows(ws: Worksheet, header_row: int, mapping: dict[str, int]) -> list[GoldRow]:
    rows: list[GoldRow] = []
    for sheet_row in range(header_row + 1, ws.max_row + 1):
        name = clean_string(_cell(ws, sheet_row, mapping.get("name")))
        if not name:
            continue
        rows.append(
            GoldRow(
                sheet_row=sheet_row,
                row_num=_as_int(_cell(ws, sheet_row, mapping.get("row_num"))),
                source_file=clean_string(_cell(ws, sheet_row, mapping.get("source_file"))),
                name=name,
                article=clean_string(_cell(ws, sheet_row, mapping.get("article"))),
                manufacturer=clean_string(_cell(ws, sheet_row, mapping.get("manufacturer"))),
                attributes=clean_string(_cell(ws, sheet_row, mapping.get("attributes"))),
                quantity=clean_string(_cell(ws, sheet_row, mapping.get("quantity"))),
                unit=clean_string(_cell(ws, sheet_row, mapping.get("unit"))),
                expected_article=clean_string(
                    _cell(ws, sheet_row, mapping.get("expected_article"))
                ),
                expected_code_1c=clean_string(
                    _cell(ws, sheet_row, mapping.get("expected_code_1c"))
                ),
                expected_name=clean_string(_cell(ws, sheet_row, mapping.get("expected_name"))),
                label_status=(
                    clean_string(_cell(ws, sheet_row, mapping.get("label_status"))) or ""
                ).lower(),
                labeler_notes=clean_string(_cell(ws, sheet_row, mapping.get("labeler_notes"))),
            )
        )
    return rows


def _cell(ws: Worksheet, row: int, col: int | None) -> Any:
    if col is None:
        return None
    return ws.cell(row=row, column=col).value


def _as_int(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


# --- Прогон матчера ---


def build_match_input(row: GoldRow) -> MatchInput:
    """Превращает строку золотого датасета в вход матчера."""
    name = row.name
    if row.attributes:
        # Характеристики из шаблона — добавим к наименованию, чтобы
        # лексический поиск имел шанс зацепиться. Это всё ещё Phase 1
        # (без извлечения структурированных атрибутов).
        name = f"{name} {row.attributes}"

    return MatchInput(
        line_number=row.sheet_row,
        name=name,
        name_normalized=normalize_name(name),
        article=row.article,
        article_normalized=normalize_article(row.article),
        manufacturer=row.manufacturer,
        manufacturer_normalized=row.manufacturer.lower() if row.manufacturer else None,
        unit=row.unit,
        unit_normalized=row.unit.lower() if row.unit else None,
    )


async def run_matcher(rows: list[GoldRow], top_k: int) -> list[RowOutcome]:
    """Прогоняет все строки через MatchingEngine."""
    settings = get_settings()
    engine = create_async_engine(
        settings.database_url_str,
        pool_pre_ping=True,
    )
    try:
        factory = async_sessionmaker(engine, expire_on_commit=False)
        async with factory() as session:
            matcher = MatchingEngine(PgTrgmSearchRepository(session))
            inputs = [build_match_input(r) for r in rows]
            results = await matcher.match_many(inputs, top_n=top_k)
    finally:
        await engine.dispose()

    outcomes: list[RowOutcome] = []
    for row, result in zip(rows, results, strict=True):
        top_articles = [normalize_article(c.article) or "" for c in result.catalog]
        top_codes = [c.code_1c or "" for c in result.catalog]

        expected_art = row.expected_article_normalized
        expected_code = (row.expected_code_1c or "").strip()

        rank = _first_match_rank(top_articles, top_codes, expected_art, expected_code)
        matched = rank == 1

        outcomes.append(
            RowOutcome(
                row=row,
                match_result=result,
                top_articles_normalized=top_articles,
                top_codes_1c=top_codes,
                expected_rank=rank,
                matched=matched,
            )
        )
    return outcomes


def _first_match_rank(
    top_articles: list[str],
    top_codes: list[str],
    expected_article: str | None,
    expected_code: str,
) -> int | None:
    """Возвращает 1-based ранг первого попадания по любому из identifier'ов.

    Совпадение засчитывается если на позиции K либо article_normalized совпал
    с expected_article (нормализованным), либо code_1c совпал с expected_code.
    """
    if not expected_article and not expected_code:
        return None
    for idx in range(len(top_articles)):
        if expected_article and top_articles[idx] == expected_article:
            return idx + 1
        if expected_code and idx < len(top_codes) and top_codes[idx] == expected_code:
            return idx + 1
    return None


# --- Метрики ---


def compute_metrics(
    outcomes: list[RowOutcome],
    *,
    top_k: int,
    min_confidence: float,
) -> GoldMetrics:
    metrics = GoldMetrics(top_k=top_k)
    metrics.total_rows = len(outcomes)

    for outcome in outcomes:
        status = outcome.row.label_status
        metrics.status_counts[status] = metrics.status_counts.get(status, 0) + 1

        if status in STATUS_UNSURE:
            metrics.skipped_unsure += 1
            continue

        if status in STATUS_FOUND or status in STATUS_ANALOG:
            if not outcome.row.has_expected_identifier:
                # Размечено как «найдено», но ни Артикул, ни Код 1С не указан —
                # такую строку нельзя соотнести с каталогом, пропускаем
                metrics.skipped_unsure += 1
                continue
            metrics.applicable += 1
            if outcome.expected_rank is not None:
                metrics.recall_at_k_hits += 1
                metrics.reciprocal_ranks_sum += 1.0 / outcome.expected_rank
                if outcome.expected_rank == 1:
                    metrics.precision_at_1_hits += 1
        elif status in STATUS_NOT_FOUND:
            metrics.not_found_rows += 1
            # Корректный «не найдено» — матчер тоже не вернул сильного кандидата
            top_cand = outcome.match_result.catalog[0] if outcome.match_result.catalog else None
            if top_cand is None or top_cand.confidence < min_confidence:
                metrics.not_found_correct += 1

    return metrics


# --- Запись результата ---


def write_results(ws: Worksheet, mapping: dict[str, int], outcomes: list[RowOutcome]) -> None:
    """Заполняет колонки M-O (Результат: артикул/уверенность/Совпало) для каждой строки."""
    article_col = mapping.get("result_article")
    code_col = mapping.get("result_code_1c")
    confidence_col = mapping.get("result_confidence")
    matched_col = mapping.get("result_matched")

    for outcome in outcomes:
        top = outcome.match_result.catalog[0] if outcome.match_result.catalog else None
        if article_col:
            ws.cell(
                row=outcome.row.sheet_row,
                column=article_col,
                value=top.article if top else None,
            )
        if code_col:
            ws.cell(
                row=outcome.row.sheet_row,
                column=code_col,
                value=top.code_1c if top else None,
            )
        if confidence_col:
            ws.cell(
                row=outcome.row.sheet_row,
                column=confidence_col,
                value=float(top.confidence) if top else None,
            )
        if matched_col:
            if not outcome.row.has_expected_identifier:
                cell_value = ""
            else:
                cell_value = "да" if outcome.matched else "нет"
            ws.cell(
                row=outcome.row.sheet_row,
                column=matched_col,
                value=cell_value,
            )


def write_metrics_sheet(wb: openpyxl.Workbook, metrics: GoldMetrics) -> None:
    """Создаёт/перезаписывает лист «Метрики» со сводкой."""
    sheet_name = "Метрики"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    ws.cell(row=1, column=1, value="Метрика").font = bold
    ws.cell(row=1, column=2, value="Значение").font = bold
    ws.cell(row=1, column=3, value="Комментарий").font = bold
    for col in (1, 2, 3):
        ws.cell(row=1, column=col).fill = header_fill

    rows = [
        ("Прогон выполнен", datetime.now(UTC).isoformat(timespec="seconds"), ""),
        ("Top-K", metrics.top_k, "глубина рассмотрения кандидатов"),
        ("", "", ""),
        ("Всего строк в датасете", metrics.total_rows, ""),
        ("Пропущено (без наименования)", metrics.skipped_no_name, ""),
        ("Пропущено (сомнительно)", metrics.skipped_unsure, ""),
        (
            "Из них применимо к метрикам",
            metrics.applicable,
            "статус найдено/аналог + есть expected",
        ),
        ("", "", ""),
        ("Recall@K", round(metrics.recall_at_k, 4), "доля строк, где правильный ответ в топ-K"),
        (
            "Precision@1",
            round(metrics.precision_at_1, 4),
            "доля строк, где правильный ответ на 1-м месте",
        ),
        ("MRR", round(metrics.mrr, 4), "среднее обратное место правильного ответа"),
        ("", "", ""),
        ("«Не найдено» в разметке", metrics.not_found_rows, ""),
        ("Из них правильно", metrics.not_found_correct, "матчер тоже не дал уверенного кандидата"),
        ("Not-found precision", round(metrics.not_found_precision, 4), ""),
    ]

    if metrics.status_counts:
        rows.append(("", "", ""))
        rows.append(("Распределение по статусам", "", ""))
        for status, count in sorted(metrics.status_counts.items()):
            label = status or "(пусто)"
            rows.append((f"  {label}", count, ""))

    for r_idx, (metric, value, comment) in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=metric)
        ws.cell(row=r_idx, column=2, value=value)
        ws.cell(row=r_idx, column=3, value=comment)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 60


def print_metrics(metrics: GoldMetrics) -> None:
    """Человекочитаемая сводка в stdout."""
    print()
    print("=" * 60)
    print(" ПРОГОН ПО ЗОЛОТОМУ ДАТАСЕТУ — РЕЗУЛЬТАТ")
    print("=" * 60)
    print(f"  Всего строк:       {metrics.total_rows}")
    print(f"  Пропущено:         {metrics.skipped_no_name + metrics.skipped_unsure}")
    print(f"  В расчёте метрик:  {metrics.applicable}")
    if metrics.applicable:
        print(f"  Recall@{metrics.top_k}:          {metrics.recall_at_k:.1%}")
        print(f"  Precision@1:       {metrics.precision_at_1:.1%}")
        print(f"  MRR:               {metrics.mrr:.3f}")
    if metrics.not_found_rows:
        print()
        print(f"  «Не найдено» rows: {metrics.not_found_rows}")
        print(
            f"  Корректно:         {metrics.not_found_correct} ({metrics.not_found_precision:.1%})"
        )
    print("=" * 60)


# --- Точка входа ---


class GoldEvalError(Exception):
    """Не удалось прочитать или обработать золотой датасет."""


def _default_output_path(input_path: Path) -> Path:
    ts = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
    return input_path.with_name(f"{input_path.stem}_results_{ts}.xlsx")


async def run_eval(
    input_path: Path,
    output_path: Path,
    *,
    sheet_name: str = "Датасет",
    top_k: int = 5,
    min_confidence: float | None = None,
) -> GoldMetrics:
    """Программная точка входа (используется и тестами, и CLI)."""
    if not input_path.exists():
        raise GoldEvalError(f"Файл не найден: {input_path}")

    settings = get_settings()
    effective_min = min_confidence if min_confidence is not None else settings.confidence_min

    wb = openpyxl.load_workbook(input_path)
    if sheet_name not in wb.sheetnames:
        raise GoldEvalError(f"Лист {sheet_name!r} не найден. Доступные: {wb.sheetnames}")
    ws = wb[sheet_name]

    header_row, mapping = detect_columns(ws)
    rows = read_rows(ws, header_row, mapping)

    if not rows:
        raise GoldEvalError("В датасете нет строк с заполненным наименованием")

    outcomes = await run_matcher(rows, top_k=top_k)
    metrics = compute_metrics(outcomes, top_k=top_k, min_confidence=effective_min)

    write_results(ws, mapping, outcomes)
    write_metrics_sheet(wb, metrics)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()

    return metrics


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="fasttender.scripts.eval_gold",
        description=(
            "Прогон матчера по золотому датасету. Заполняет колонки результата "
            "и считает Recall@K / Precision@1 / MRR (раздел 20.2)."
        ),
    )
    parser.add_argument("input", type=Path, help="Путь к gold_dataset_template.xlsx")
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        default=None,
        help="Путь выходного файла (по умолчанию рядом с входным)",
    )
    parser.add_argument(
        "--top-k",
        type=int,
        default=5,
        help="Глубина рассмотрения топ-кандидатов (по умолчанию 5)",
    )
    parser.add_argument(
        "--sheet",
        default="Датасет",
        help='Имя листа с данными (по умолчанию "Датасет")',
    )
    parser.add_argument(
        "--min-confidence",
        type=float,
        default=None,
        help="Порог уверенности для not-found-метрики (по умолчанию из настроек)",
    )

    args = parser.parse_args(argv)

    input_path: Path = args.input.resolve()
    output_path: Path = (args.output or _default_output_path(input_path)).resolve()

    try:
        metrics = asyncio.run(
            run_eval(
                input_path=input_path,
                output_path=output_path,
                sheet_name=args.sheet,
                top_k=args.top_k,
                min_confidence=args.min_confidence,
            )
        )
    except GoldEvalError as exc:
        print(f"Ошибка: {exc}", file=sys.stderr)
        return 2

    print_metrics(metrics)
    print(f"\nРезультат сохранён в: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
