"""Экспорт золотого датасета в Excel-шаблон (Приложение C.3).

Формирует лист «Датасет» с двухуровневой шапкой ровно в том формате, который
читает `eval_gold.py` (групповые лейблы в R1, заголовки в R2, данные с R3,
17 колонок). Колонки результата (14–17) оставляем пустыми — их заполняет
прогон матчера. Так связка «веб-ввод → экспорт → CLI eval_gold» работает без
изменений в самом прогоне.
"""

from __future__ import annotations

import io

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.cell import get_column_letter

from fasttender.models import GoldRow

# Заголовки строки R2 — порядок повторяет реальный шаблон
# gold_dataset_template_new.xlsx (см. tests/integration/test_eval_gold.py).
GOLD_HEADERS: tuple[str, ...] = (
    "№",
    "Файл-источник",
    "Наименование (как у клиента)",
    "Артикул (как у клиента)",
    "Производитель (как у клиента)",
    "Характеристика (как у клиента)",
    "Кол-во",
    "Ед. изм.",
    "→ Правильный артикул каталога",
    "→ Код 1С каталога",
    "→ Правильное наименование каталога",
    "Статус разметки",
    "Примечание разметчика",
    "Результат матчера: артикул",
    "Результат матчера: Код 1С",
    "Результат матчера: уверенность",
    "Совпало? (да/нет)",
)

SHEET_NAME = "Датасет"


def _quantity_value(row: GoldRow) -> float | int | None:
    if row.quantity is None:
        return None
    q = float(row.quantity)
    return int(q) if q.is_integer() else q


def build_gold_xlsx(rows: list[GoldRow]) -> bytes:
    """Собирает XLSX золотого датасета в формате шаблона. Возвращает байты."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    # R1 — групповые лейблы (для человека; парсер их не читает)
    ws.cell(row=1, column=2, value="ИСХОДНЫЕ ДАННЫЕ (как в спецификации клиента)").font = bold
    ws.cell(row=1, column=9, value="ЭТАЛОННАЯ РАЗМЕТКА").font = bold
    ws.cell(row=1, column=14, value="РЕЗУЛЬТАТ ПРОГОНА").font = bold

    # R2 — заголовки колонок
    for c_idx, header in enumerate(GOLD_HEADERS, start=1):
        cell = ws.cell(row=2, column=c_idx, value=header)
        cell.font = bold
        cell.fill = header_fill

    # Данные с R3. Колонки результата (14–17) оставляем пустыми.
    for offset, row in enumerate(rows):
        r = 3 + offset
        ws.cell(row=r, column=1, value=offset + 1)
        ws.cell(row=r, column=2, value=row.source_file)
        ws.cell(row=r, column=3, value=row.name)
        ws.cell(row=r, column=4, value=row.article)
        ws.cell(row=r, column=5, value=row.manufacturer)
        ws.cell(row=r, column=6, value=row.attributes)
        ws.cell(row=r, column=7, value=_quantity_value(row))
        ws.cell(row=r, column=8, value=row.unit)
        ws.cell(row=r, column=9, value=row.expected_article)
        ws.cell(row=r, column=10, value=row.expected_code_1c)
        ws.cell(row=r, column=11, value=row.expected_name)
        ws.cell(row=r, column=12, value=row.label_status.value)
        ws.cell(row=r, column=13, value=row.labeler_notes)

    for c_idx in range(1, len(GOLD_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()
