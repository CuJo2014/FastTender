"""Парсеры Excel-файлов (XLSX/XLSM через openpyxl, XLS через xlrd).

Оба формата разворачиваются в общий промежуточный вид `list[list[Any]]` —
плоская таблица одного листа с развёрнутыми объединёнными ячейками
(раздел 4.1.4). Дальше — общий код в `_matrix.build_result`.
"""

from pathlib import Path
from typing import Any

import openpyxl
import xlrd
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from fasttender.services.parser._matrix import build_result
from fasttender.services.parser.types import ColumnMapping, ParseError, ParseResult, SpecField


def parse_excel(
    path: Path,
    *,
    sheet_name: str | None = None,
    mapping_override: ColumnMapping | None = None,
    exclude_fields: frozenset[SpecField] | None = None,
) -> ParseResult:
    """Парсит Excel-файл (.xlsx/.xlsm/.xls) и возвращает ParseResult."""
    ext = path.suffix.lower()
    if ext in {".xlsx", ".xlsm"}:
        title, matrix = _read_xlsx_sheet(path, sheet_name)
    elif ext == ".xls":
        title, matrix = _read_xls_sheet(path, sheet_name)
    else:
        raise ParseError(f"Неподдерживаемое расширение Excel: {ext}")

    return build_result(
        matrix,
        sheet_name=title,
        mapping_override=mapping_override,
        exclude_fields=exclude_fields,
    )


def _read_xlsx_sheet(path: Path, sheet_name: str | None) -> tuple[str, list[list[Any]]]:
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:
        raise ParseError(
            f"Не удалось открыть XLSX-файл: {exc}",
            details={"path": str(path)},
        ) from exc

    try:
        ws: Worksheet
        if sheet_name is not None:
            if sheet_name not in wb.sheetnames:
                raise ParseError(
                    f"Лист {sheet_name!r} не найден",
                    details={"available": wb.sheetnames},
                )
            ws = wb[sheet_name]
        else:
            ws = wb.worksheets[0]
            for candidate in wb.worksheets:
                if candidate.max_row > 1:
                    ws = candidate
                    break

        rows = _materialize_worksheet(ws)
        return ws.title, rows
    finally:
        wb.close()


def _materialize_worksheet(ws: Worksheet) -> list[list[Any]]:
    """Превращает worksheet в плотный list[list] и разворачивает merged-ячейки.

    Для merged-диапазонов значение из верхней-левой ячейки копируется во все
    остальные ячейки диапазона — стандартный приём для парсинга «грязных» таблиц.
    """
    matrix: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        matrix.append(list(row))

    if not matrix:
        return matrix

    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        anchor_value = matrix[min_row - 1][min_col - 1]
        for r in range(min_row - 1, max_row):
            while len(matrix[r]) < max_col:
                matrix[r].append(None)
            for c in range(min_col - 1, max_col):
                matrix[r][c] = anchor_value

    return matrix


def _read_xls_sheet(path: Path, sheet_name: str | None) -> tuple[str, list[list[Any]]]:
    try:
        book = xlrd.open_workbook(str(path), formatting_info=False)
    except Exception as exc:
        raise ParseError(
            f"Не удалось открыть XLS-файл: {exc}",
            details={"path": str(path)},
        ) from exc

    if sheet_name is not None:
        try:
            sheet = book.sheet_by_name(sheet_name)
        except xlrd.XLRDError as exc:
            raise ParseError(
                f"Лист {sheet_name!r} не найден",
                details={"available": book.sheet_names()},
            ) from exc
    else:
        sheet = book.sheet_by_index(0)
        for i in range(book.nsheets):
            candidate = book.sheet_by_index(i)
            if candidate.nrows > 1:
                sheet = candidate
                break

    matrix: list[list[Any]] = []
    for r in range(sheet.nrows):
        row: list[Any] = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            row.append(cell.value if cell.value != "" else None)
        matrix.append(row)
    return sheet.name, matrix
