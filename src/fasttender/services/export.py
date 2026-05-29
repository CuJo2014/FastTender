"""Экспорт результатов обработки спецификации (раздел 4.8).

Phase 1: XLSX и CSV. В Фазе 2 добавится генерация КП по фирменному
шаблону (DOCX/PDF) и JSON для интеграций.

Колонки одинаковые для обоих форматов:
  №, Наименование (как у клиента), Артикул, Производитель, Кол-во,
  Ед.изм., Цена клиента, ────, Решение, Источник, Артикул выбранной,
  Наименование выбранной, Производитель выбранной, Цена, Валюта,
  Ед.изм. выбранной, Confidence, Тип совпадения, Примечание
"""

import csv
import io
from datetime import UTC, datetime
from enum import StrEnum
from typing import Any
from uuid import UUID

import openpyxl
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from fasttender.models import (
    DataSource,
    DataSourceType,
    Item,
    MatchCandidate,
    Specification,
    SpecItem,
    Supplier,
    Verification,
    VerificationDecision,
)


class ExportFormat(StrEnum):
    XLSX = "xlsx"
    CSV = "csv"


HEADERS = (
    "№",
    "Наименование (клиент)",
    "Артикул (клиент)",
    "Производитель (клиент)",
    "Кол-во",
    "Ед. изм.",
    "Цена клиента",
    "Решение",
    "Источник",
    "Категория каталога",
    "Артикул выбранной",
    "Код 1С",
    "Наименование выбранной",
    "Производитель выбранной",
    "Цена",
    "Валюта",
    "Ед. изм. выбранной",
    "Confidence",
    "Тип совпадения",
    "Примечание",
)


async def build_export(
    session: AsyncSession,
    spec: Specification,
    fmt: ExportFormat,
) -> tuple[bytes, str, str]:
    """Возвращает (содержимое, content_type, имя_файла)."""
    rows = await _collect_rows(session, spec.id)

    timestamp = datetime.now(UTC).strftime("%Y%m%d-%H%M")
    base_name = _safe_filename(spec.source_filename or f"spec-{spec.id}")
    filename = f"{base_name}_{timestamp}.{fmt.value}"

    if fmt is ExportFormat.XLSX:
        content = _to_xlsx(rows, spec)
        return (
            content,
            ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            filename,
        )

    if fmt is ExportFormat.CSV:
        content = _to_csv(rows)
        return content, "text/csv; charset=utf-8", filename

    raise ValueError(f"Неизвестный формат экспорта: {fmt}")  # pragma: no cover


# --- Сбор данных ---


async def _collect_rows(session: AsyncSession, spec_id: UUID) -> list[dict[str, Any]]:
    """Готовит данные для экспорта построчно.

    Для каждой строки спецификации:
      1. Если есть Verification → выбираем chosen_item (если confirmed),
         либо помечаем «не найдено» / «требуется новый».
      2. Иначе берём топ-1 catalog (приоритетнее), затем топ-1 supplier.
      3. Если кандидатов нет — пишем строку «нет совпадения».
    """
    spec_items = (
        await session.scalars(
            select(SpecItem)
            .where(SpecItem.spec_id == spec_id)
            .order_by(SpecItem.line_number)
            .options(
                selectinload(SpecItem.candidates)
                .selectinload(MatchCandidate.item)
                .selectinload(Item.source)
                .selectinload(DataSource.supplier),
                selectinload(SpecItem.verification)
                .selectinload(Verification.chosen_item)
                .selectinload(Item.source)
                .selectinload(DataSource.supplier),
            )
        )
    ).all()

    rows: list[dict[str, Any]] = []
    for spec_item in spec_items:
        chosen_item, decision_label, notes = _decide_winner(spec_item)
        confidence = _confidence_for(spec_item, chosen_item)
        match_type_label = _match_type_for(spec_item, chosen_item)

        rows.append(
            {
                "line_number": spec_item.line_number,
                "name_raw": spec_item.name_raw,
                "article_raw": spec_item.article_raw,
                "manufacturer_raw": spec_item.manufacturer_raw,
                "quantity": _format_decimal(spec_item.quantity),
                "unit_raw": spec_item.unit_raw,
                "price_raw": _format_decimal(spec_item.price_raw),
                "decision": decision_label,
                "source": _source_label(chosen_item),
                "category_path": chosen_item.category_path if chosen_item else None,
                "chosen_article": chosen_item.article_raw if chosen_item else None,
                "chosen_code_1c": chosen_item.code_1c if chosen_item else None,
                "chosen_name": chosen_item.name if chosen_item else None,
                "chosen_manufacturer": chosen_item.manufacturer if chosen_item else None,
                "chosen_price": _format_decimal(chosen_item.price) if chosen_item else None,
                "chosen_currency": chosen_item.currency if chosen_item else None,
                "chosen_unit": chosen_item.unit if chosen_item else None,
                "confidence": f"{confidence:.3f}" if confidence is not None else None,
                "match_type": match_type_label,
                "notes": notes,
            }
        )
    return rows


def _decide_winner(spec_item: SpecItem) -> tuple[Item | None, str, str | None]:
    """Какую позицию показывать в выгрузке + текст решения и примечание.

    Возвращает (item-или-None, человекочитаемое-решение, notes).
    """
    verification = spec_item.verification
    if verification is not None:
        if verification.decision is VerificationDecision.CONFIRMED and verification.chosen_item:
            return verification.chosen_item, "Подтверждено", verification.notes
        if verification.decision is VerificationDecision.NOT_FOUND:
            return None, "Не найдено (отметка менеджера)", verification.notes
        if verification.decision is VerificationDecision.REJECTED:
            return None, "Отклонено", verification.notes
        if verification.decision is VerificationDecision.NEW_ITEM_REQUESTED:
            return None, "Требуется новая позиция каталога", verification.notes

    # Без верификации — топ-1 catalog приоритетнее supplier
    top = _pick_top_candidate(spec_item.candidates)
    if top is None:
        return None, "Не найдено (нет кандидатов)", None
    return top.item, "Не верифицировано", None


def _pick_top_candidate(candidates: list[MatchCandidate]) -> MatchCandidate | None:
    if not candidates:
        return None

    by_type: dict[DataSourceType, list[MatchCandidate]] = {}
    for c in candidates:
        by_type.setdefault(c.item.source.type, []).append(c)

    for type_ in (DataSourceType.COMPANY_CATALOG, DataSourceType.SUPPLIER_PRICELIST):
        group = by_type.get(type_)
        if group:
            return min(group, key=lambda c: c.rank)
    return None


def _confidence_for(spec_item: SpecItem, chosen_item: Item | None) -> float | None:
    if chosen_item is None:
        return None
    # Ищем MatchCandidate с этим item — он мог быть rank=2 если менеджер выбрал
    # не топ-1, но всё ещё в топ-N
    for c in spec_item.candidates:
        if c.item_id == chosen_item.id:
            return float(c.confidence)
    return None


def _match_type_for(spec_item: SpecItem, chosen_item: Item | None) -> str | None:
    if chosen_item is None:
        return None
    for c in spec_item.candidates:
        if c.item_id == chosen_item.id:
            return c.match_type.value
    return None


def _source_label(item: Item | None) -> str | None:
    if item is None:
        return None
    source = item.source
    if source.type is DataSourceType.COMPANY_CATALOG:
        return "Каталог компании"
    supplier: Supplier | None = source.supplier
    if supplier is not None:
        return f"Прайс: {supplier.name}"
    return source.name


def _format_decimal(value: Any) -> str | None:
    if value is None:
        return None
    # Decimal без лишних нулей для целых
    text = f"{value:f}".rstrip("0").rstrip(".") if "." in f"{value:f}" else f"{value}"
    return text or "0"


def _safe_filename(name: str) -> str:
    """Убирает расширение и кириллицу-небезопасные символы для HTTP-заголовка."""
    base = name.rsplit(".", 1)[0] if "." in name else name
    return "".join(c if c.isalnum() or c in "._- " else "_" for c in base) or "spec"


# --- Сериализация ---


def _to_xlsx(rows: list[dict[str, Any]], spec: Specification) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результат"

    # Заголовок-метаданные
    ws.append([f"Спецификация: {spec.source_filename}"])
    if spec.client_name:
        ws.append([f"Клиент: {spec.client_name}"])
    ws.append([f"Сформировано: {datetime.now(UTC).isoformat(timespec='seconds')}"])
    ws.append([])

    # Шапка таблицы
    header_row_idx = ws.max_row + 1
    ws.append(list(HEADERS))
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    # Данные
    for row in rows:
        ws.append(_row_to_tuple(row))

    # Авто-ширина (грубо)
    widths = [
        4,  # №
        35,  # Наименование (клиент)
        18,  # Артикул (клиент)
        18,  # Производитель (клиент)
        8,  # Кол-во
        8,  # Ед. изм.
        12,  # Цена клиента
        18,  # Решение
        22,  # Источник
        28,  # Категория каталога
        18,  # Артикул выбранной
        16,  # Код 1С
        35,  # Наименование выбранной
        18,  # Производитель выбранной
        12,  # Цена
        8,  # Валюта
        8,  # Ед. изм. выбранной
        10,  # Confidence
        14,  # Тип совпадения
        30,  # Примечание
    ]
    from openpyxl.utils import get_column_letter

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _to_csv(rows: list[dict[str, Any]]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(HEADERS)
    for row in rows:
        writer.writerow(["" if v is None else str(v) for v in _row_to_tuple(row)])

    # BOM для совместимости с Excel при открытии CSV в кириллице
    return "﻿".encode() + buf.getvalue().encode("utf-8")


def _row_to_tuple(row: dict[str, Any]) -> tuple[Any, ...]:
    return (
        row["line_number"],
        row["name_raw"],
        row["article_raw"],
        row["manufacturer_raw"],
        row["quantity"],
        row["unit_raw"],
        row["price_raw"],
        row["decision"],
        row["source"],
        row["category_path"],
        row["chosen_article"],
        row["chosen_code_1c"],
        row["chosen_name"],
        row["chosen_manufacturer"],
        row["chosen_price"],
        row["chosen_currency"],
        row["chosen_unit"],
        row["confidence"],
        row["match_type"],
        row["notes"],
    )
