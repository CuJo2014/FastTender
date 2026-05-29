"""Интеграционные тесты экспорта XLSX/CSV."""

import io
from pathlib import Path

import openpyxl
import pytest
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from fasttender.models import (
    Item,
    Specification,
    SpecificationStatus,
    SpecItem,
)
from fasttender.models.enums import VerificationDecision
from fasttender.services.export import HEADERS, ExportFormat, build_export
from fasttender.services.importer import CatalogImporter, ImportMode
from fasttender.services.pipeline import SpecificationProcessor
from fasttender.services.verification import VerificationService
from tests.fixtures.spec_builders import make_xlsx


async def _setup_spec_with_results(session: AsyncSession, tmp_path: Path) -> Specification:
    catalog = make_xlsx(
        tmp_path / "catalog.xlsx",
        rows=[
            ["Артикул", "Наименование", "Производитель", "Ед.", "Цена"],
            ["BLT-001", "Болт М10х40 DIN933", "KOELNER", "шт", "10.00"],
            ["NUT-001", "Гайка М10 DIN934", "KOELNER", "шт", "4.00"],
        ],
    )
    await CatalogImporter().import_file(session, catalog, mode=ImportMode.REPLACE)
    await session.commit()

    spec_file = make_xlsx(
        tmp_path / "spec.xlsx",
        rows=[
            ["Наименование", "Артикул", "Кол-во", "Ед."],
            ["Болт М10х40", "BLT-001", 50, "шт"],
            ["Гайка М10", "NUT-001", 100, "шт"],
            ["Совсем непонятная позиция xyz", "ZZZ-NOPE", 1, "шт"],
        ],
    )
    spec = Specification(
        source_filename="spec.xlsx",
        storage_path=str(spec_file),
        client_name="ООО Ромашка",
        status=SpecificationStatus.UPLOADED,
        meta={},
    )
    session.add(spec)
    await session.commit()
    await session.refresh(spec)

    await SpecificationProcessor(session).process(spec.id)
    return spec


# --- XLSX ---


async def test_xlsx_export_has_headers_and_data(session: AsyncSession, tmp_path: Path) -> None:
    spec = await _setup_spec_with_results(session, tmp_path)
    content, content_type, filename = await build_export(session, spec, ExportFormat.XLSX)

    assert content_type.endswith("spreadsheetml.sheet")
    assert filename.startswith("spec_")
    assert filename.endswith(".xlsx")

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    # Метаданные в первых строках
    rows = list(ws.iter_rows(values_only=True))
    assert any("ООО Ромашка" in (r[0] or "") for r in rows[:5])

    # Шапка
    header_row = next(r for r in rows if r[0] == HEADERS[0])
    assert tuple(header_row[: len(HEADERS)]) == HEADERS

    # Данные — все 3 строки спецификации
    data_rows = rows[rows.index(header_row) + 1 :]
    assert len(data_rows) == 3

    # Первая — Болт, не верифицирована, но топ-1 catalog взят
    line1 = data_rows[0]
    assert line1[1] == "Болт М10х40"  # name_raw
    assert line1[2] == "BLT-001"  # article_raw
    assert line1[7] == "Не верифицировано"  # decision
    assert "Каталог" in line1[8]  # source label
    # line1[9] = category_path; в seed-каталоге без категории → None
    assert line1[9] is None
    assert line1[11] == "Болт М10х40 DIN933"  # chosen_name (index 11 после Категории)
    assert float(line1[16]) >= 0.95  # confidence (index 16)

    # Третья — ZZZ-NOPE, нет catalog кандидатов
    line3 = data_rows[2]
    assert line3[7] == "Не найдено (нет кандидатов)"

    wb.close()


async def test_xlsx_export_uses_verification_when_present(
    session: AsyncSession, tmp_path: Path
) -> None:
    spec = await _setup_spec_with_results(session, tmp_path)

    # Подтверждаем первую строку явно
    blt = await session.scalar(
        select(SpecItem).where(SpecItem.spec_id == spec.id, SpecItem.article_normalized == "BLT001")
    )
    nut_item = await session.scalar(  # выбираем гайку для болта — нарочно «не тот» товар
        select(Item).where(Item.article_normalized == "NUT001")
    )
    await VerificationService(session).upsert(
        spec_id=spec.id,
        spec_item_id=blt.id,
        decision=VerificationDecision.CONFIRMED,
        chosen_item_id=nut_item.id,
        notes="Заменили на гайку",
    )
    await session.commit()

    content, _, _ = await build_export(session, spec, ExportFormat.XLSX)
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header_row = next(r for r in rows if r[0] == HEADERS[0])
    data_rows = rows[rows.index(header_row) + 1 :]
    line1 = data_rows[0]
    assert line1[7] == "Подтверждено"
    # Менеджер выбрал гайку, не болт. Индексы сдвинулись на 1 после
    # вставки колонки «Категория каталога» в позицию 9.
    assert line1[11] == "Гайка М10 DIN934"  # chosen_name
    assert line1[18] == "Заменили на гайку"  # notes
    wb.close()


async def test_xlsx_export_marks_not_found_for_rejected(
    session: AsyncSession, tmp_path: Path
) -> None:
    spec = await _setup_spec_with_results(session, tmp_path)
    unknown = await session.scalar(
        select(SpecItem).where(
            SpecItem.spec_id == spec.id, SpecItem.article_normalized == "ZZZNOPE"
        )
    )
    await VerificationService(session).upsert(
        spec_id=spec.id,
        spec_item_id=unknown.id,
        decision=VerificationDecision.NOT_FOUND,
    )
    await session.commit()

    content, _, _ = await build_export(session, spec, ExportFormat.XLSX)
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    header_row = next(r for r in rows if r[0] == HEADERS[0])
    data_rows = rows[rows.index(header_row) + 1 :]
    line3 = data_rows[2]
    assert line3[7] == "Не найдено (отметка менеджера)"
    assert line3[11] is None  # chosen_name пуст (index сдвинулся после вставки Категории)
    wb.close()


# --- CSV ---


async def test_csv_export_has_bom_and_semicolon(session: AsyncSession, tmp_path: Path) -> None:
    spec = await _setup_spec_with_results(session, tmp_path)
    content, content_type, filename = await build_export(session, spec, ExportFormat.CSV)

    assert content_type.startswith("text/csv")
    assert filename.endswith(".csv")

    # UTF-8 BOM
    assert content.startswith(b"\xef\xbb\xbf")
    text = content.decode("utf-8")
    # Excel-friendly разделитель
    assert ";" in text
    # Шапка
    assert "Наименование (клиент)" in text
    # Все 3 позиции
    assert text.count("Болт М10х40") >= 1
    assert "ZZZ" in text or "Совсем непонятная" in text


@pytest.mark.parametrize("fmt", [ExportFormat.XLSX, ExportFormat.CSV])
async def test_export_filename_is_safe(
    session: AsyncSession, tmp_path: Path, fmt: ExportFormat
) -> None:
    spec = await _setup_spec_with_results(session, tmp_path)
    spec.source_filename = "Spec/with*bad:chars?.xlsx"
    await session.commit()
    _, _, filename = await build_export(session, spec, fmt)
    # Без слешей и звёздочек
    assert "/" not in filename
    assert "*" not in filename
    assert ":" not in filename
    assert "?" not in filename
