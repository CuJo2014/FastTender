"""End-to-end проверка кода 1С: парсер → импорт → API → eval matching by code.

См. обсуждение 2026-05-29: 1С хранит два разных идентификатора —
Артикул (артикул производителя) и Код (внутренний 1С ID). Phase 1
разделяет их на разные поля Item.article_raw и Item.code_1c.
"""

from collections.abc import AsyncIterator
from pathlib import Path

import openpyxl
import pytest
import pytest_asyncio
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from fasttender.models import Item
from fasttender.scripts.eval_gold import detect_columns, run_eval
from fasttender.services.importer import CatalogImporter, ImportMode
from fasttender.services.parser import SpecField, SpecificationParser
from tests.fixtures.spec_builders import make_xlsx
from tests.integration.conftest import TEST_DB_URL
from tests.integration.test_eval_gold import make_gold_xlsx

# Локальная committed_db: для тестов, где run_eval создаёт свой engine
# и должен видеть commits.
_TABLES = (
    "verification",
    "match_candidate",
    "spec_item",
    "specification",
    "item",
    "data_source",
    "supplier",
)


@pytest_asyncio.fixture
async def committed_db() -> AsyncIterator[AsyncSession]:
    engine = create_async_engine(TEST_DB_URL, future=True)
    async with engine.connect() as connection:
        await connection.execute(text(f"TRUNCATE {', '.join(_TABLES)} RESTART IDENTITY CASCADE"))
        await connection.commit()

    factory = async_sessionmaker(engine, expire_on_commit=False)
    async with factory() as s:
        yield s

    async with engine.connect() as connection:
        await connection.execute(text(f"TRUNCATE {', '.join(_TABLES)} RESTART IDENTITY CASCADE"))
        await connection.commit()
    await engine.dispose()


def test_parser_distinguishes_article_from_code(tmp_path: Path) -> None:
    """Колонки «Артикул» и «Код» — два разных поля в ParsedItem."""
    parser = SpecificationParser()
    path = make_xlsx(
        tmp_path / "catalog.xlsx",
        rows=[
            ["Артикул", "Код", "Наименование", "Цена"],
            ["BLT-001", "Ц0000000123", "Болт М10", "10"],
            [None, "Ц0000000124", "Гайка без артикула", "4"],
        ],
    )
    result = parser.parse(path)
    assert result.column_mapping.has(SpecField.ARTICLE)
    assert result.column_mapping.has(SpecField.CODE_1C)

    assert result.items[0].article == "BLT-001"
    assert result.items[0].code_1c == "Ц0000000123"
    assert result.items[1].article is None
    assert result.items[1].code_1c == "Ц0000000124"


async def test_catalog_import_stores_both_identifiers(
    session: AsyncSession,
    tmp_path: Path,
) -> None:
    catalog = make_xlsx(
        tmp_path / "cat.xlsx",
        rows=[
            ["Артикул", "Код", "Наименование", "Цена"],
            ["BLT-001", "Ц0000000123", "Болт М10", "10"],
            [None, "Ц0000000124", "Гайка без артикула", "4"],
            ["ABC-9", "Ц0000000125", "Третья", "5"],
        ],
    )
    await CatalogImporter().import_file(session, catalog, mode=ImportMode.REPLACE)
    await session.commit()

    items = {i.code_1c: i for i in (await session.scalars(select(Item))).all()}
    assert items["Ц0000000123"].article_raw == "BLT-001"
    assert items["Ц0000000124"].article_raw is None
    assert items["Ц0000000125"].article_raw == "ABC-9"


async def test_unique_constraint_per_active_code(
    session: AsyncSession,
    tmp_path: Path,
) -> None:
    """Один Код 1С не может дублироваться в одном активном источнике."""
    catalog = make_xlsx(
        tmp_path / "cat.xlsx",
        rows=[
            ["Артикул", "Код", "Наименование", "Цена"],
            ["A-1", "Ц0000000100", "Первая", "1"],
            ["A-2", "Ц0000000200", "Вторая", "2"],
        ],
    )
    await CatalogImporter().import_file(session, catalog, mode=ImportMode.REPLACE)
    await session.commit()

    # Повторный REPLACE того же файла — миграция 0005 (как и 0004) должна это переживать
    await CatalogImporter().import_file(session, catalog, mode=ImportMode.REPLACE)
    await session.commit()

    actives = (await session.scalars(select(Item).where(Item.is_active))).all()
    codes = [i.code_1c for i in actives]
    assert sorted(codes) == ["Ц0000000100", "Ц0000000200"]


async def test_eval_matches_by_code_1c_when_no_article(
    committed_db: AsyncSession,
    tmp_path: Path,
) -> None:
    """Gold dataset с expected_code_1c (без article) — eval ищет совпадение по Коду."""
    catalog = make_xlsx(
        tmp_path / "cat.xlsx",
        rows=[
            ["Артикул", "Код", "Наименование", "Цена"],
            [None, "Ц0000000500", "Снегоуборщик HUTER SGC 8100", "50000"],
        ],
    )
    await CatalogImporter().import_file(committed_db, catalog, mode=ImportMode.REPLACE)
    await committed_db.commit()

    # Gold dataset с expected_code_1c, без expected_article
    gold_path = make_gold_xlsx(
        tmp_path / "gold.xlsx",
        data_rows=[
            (
                1,  # № (row_num)
                "spec.xlsx",  # source_file
                "Снегоуборщик HUTER",  # name
                None,  # article (клиент не написал)
                "HUTER",  # manufacturer
                None,  # attributes
                1,  # quantity
                "шт",  # unit
                None,  # expected_article (нет)
                "Ц0000000500",  # expected_code_1c
                "Снегоуборщик HUTER SGC 8100",  # expected_name
                "найдено",  # label_status
                "матчим по коду 1С",  # labeler_notes
            ),
        ],
    )
    output = tmp_path / "result.xlsx"
    metrics = await run_eval(input_path=gold_path, output_path=output, top_k=5)

    assert metrics.applicable == 1
    assert metrics.recall_at_k_hits == 1
    assert metrics.precision_at_1_hits == 1

    # Проверим что в выходном файле колонка «Результат: Код 1С» заполнена
    wb_out = openpyxl.load_workbook(output)
    ws = wb_out["Датасет"]
    _header_row, mapping = detect_columns(ws)
    assert "result_code_1c" in mapping
    code_col = mapping["result_code_1c"]
    matched_col = mapping["result_matched"]
    assert ws.cell(row=3, column=code_col).value == "Ц0000000500"
    assert ws.cell(row=3, column=matched_col).value == "да"
    wb_out.close()


async def test_eval_matches_by_article_when_present(
    committed_db: AsyncSession,
    tmp_path: Path,
) -> None:
    """Если есть Артикул — eval по нему. Код не обязателен."""
    catalog = make_xlsx(
        tmp_path / "cat.xlsx",
        rows=[
            ["Артикул", "Код", "Наименование", "Цена"],
            ["SGC8100", "Ц0000000500", "Снегоуборщик HUTER SGC 8100", "50000"],
        ],
    )
    await CatalogImporter().import_file(committed_db, catalog, mode=ImportMode.REPLACE)
    await committed_db.commit()

    gold_path = make_gold_xlsx(
        tmp_path / "gold.xlsx",
        data_rows=[
            (
                1,
                "spec.xlsx",
                "Снегоуборщик HUTER SGC 8100",
                "SGC 8100",  # article написан клиентом
                "HUTER",
                None,
                1,
                "шт",
                "SGC8100",  # expected_article (нормализованный)
                None,  # expected_code_1c (не указан)
                "Снегоуборщик HUTER SGC 8100",
                "найдено",
                None,
            ),
        ],
    )
    output = tmp_path / "result.xlsx"
    metrics = await run_eval(input_path=gold_path, output_path=output, top_k=5)

    assert metrics.applicable == 1
    assert metrics.recall_at_k_hits == 1
    assert metrics.precision_at_1_hits == 1


# Silence unused
_ = pytest
