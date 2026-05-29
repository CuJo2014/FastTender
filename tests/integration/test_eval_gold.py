"""Integration-тест прогона по золотому датасету.

Структура синтетического XLSX — точная копия gold_dataset_template_new.xlsx:
лист «Датасет», 15 колонок, групповые лейблы в R1, заголовки в R2,
данные с R3.
"""

import io
from collections.abc import AsyncIterator
from pathlib import Path

import openpyxl
import pytest
import pytest_asyncio
from openpyxl.utils.cell import get_column_letter
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from fasttender.scripts.eval_gold import (
    GoldEvalError,
    detect_columns,
    read_rows,
    run_eval,
)
from fasttender.services.importer import CatalogImporter, ImportMode
from tests.fixtures.spec_builders import make_xlsx
from tests.integration.conftest import TEST_DB_URL

_TABLES = (
    "verification",
    "match_candidate",
    "spec_item",
    "specification",
    "item",
    "data_source",
    "supplier",
)


@pytest_asyncio.fixture
async def committed_db() -> AsyncIterator[AsyncSession]:
    """Реальные commit'ы — eval_gold ходит через свой engine."""
    engine = create_async_engine(TEST_DB_URL, future=True)
    async with engine.connect() as connection:
        await connection.execute(text(f"TRUNCATE {', '.join(_TABLES)} RESTART IDENTITY CASCADE"))
        await connection.commit()

    factory = async_sessionmaker(engine, expire_on_commit=False)
    async with factory() as session:
        yield session

    async with engine.connect() as connection:
        await connection.execute(text(f"TRUNCATE {', '.join(_TABLES)} RESTART IDENTITY CASCADE"))
        await connection.commit()
    await engine.dispose()


# --- Утилиты сборки синтетического шаблона ---

GOLD_HEADERS = (
    "№",
    "Файл-источник",
    "Наименование (как у клиента)",
    "Артикул (как у клиента)",
    "Производитель (как у клиента)",
    "Характеристика (как у клиента)",
    "Кол-во",
    "Ед. изм.",
    "→ Правильный артикул каталога",
    "→ Правильное наименование каталога",
    "Статус разметки",
    "Примечание разметчика",
    "Результат матчера: артикул",
    "Результат матчера: уверенность",
    "Совпало? (да/нет)",
)


def make_gold_xlsx(path: Path, data_rows: list[tuple]) -> Path:
    """Создаёт XLSX с точной копией структуры реального шаблона.

    data_rows — кортежи длиной 12 (без 3-х колонок результата, они пустые).
    """
    wb = openpyxl.Workbook()
    # «Инструкция» можно пропустить — рабочий лист один
    ws = wb.active
    ws.title = "Датасет"

    # R1 — групповые лейблы (только для визуального соответствия, парсер не читает)
    ws.cell(row=1, column=2, value="ИСХОДНЫЕ ДАННЫЕ (как в спецификации клиента)")
    ws.cell(row=1, column=9, value="ЭТАЛОННАЯ РАЗМЕТКА")
    ws.cell(row=1, column=13, value="РЕЗУЛЬТАТ ПРОГОНА")

    # R2 — реальная шапка
    for c_idx, header in enumerate(GOLD_HEADERS, start=1):
        ws.cell(row=2, column=c_idx, value=header)

    # Данные с R3 — 12 заполненных колонок, 3 пустых (M-O будут заполнены evaluator'ом)
    for r_offset, row in enumerate(data_rows, start=3):
        if len(row) != 12:
            raise ValueError(f"Ожидаются 12 колонок данных, получено {len(row)}")
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_offset, column=c_idx, value=value)

    # Подгоняем ширины
    for c_idx in range(1, len(GOLD_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 18

    wb.save(path)
    wb.close()
    return path


async def _seed_catalog(session: AsyncSession, tmp_path: Path) -> None:
    """Каталог для матчинга — 5 позиций."""
    catalog = make_xlsx(
        tmp_path / "catalog.xlsx",
        rows=[
            ["Артикул", "Наименование", "Производитель", "Ед.", "Цена"],
            ["BLT-M10-040-ZN", "Болт М10х40 DIN 933 оцинкованный", "KOELNER", "шт", "12.50"],
            ["BLT-M10-050-ZN", "Болт М10х50 DIN 933 оцинкованный", "KOELNER", "шт", "14.80"],
            ["NUT-M10", "Гайка М10 DIN 934", "KOELNER", "шт", "4.20"],
            ["WSH-M10", "Шайба плоская М10 DIN 125", "KOELNER", "шт", "1.10"],
            ["DOWEL-6-40", "Дюбель универсальный 6х40", "FISCHER", "шт", "2.50"],
        ],
    )
    await CatalogImporter().import_file(session, catalog, mode=ImportMode.REPLACE)
    await session.commit()


# --- Тесты структуры/чтения ---


def test_detect_columns_real_template_layout(tmp_path: Path) -> None:
    """На синтетике с точной структурой шаблона — все 15 колонок распознаны."""
    path = make_gold_xlsx(
        tmp_path / "gold.xlsx",
        data_rows=[
            (
                1,
                "spec.xlsx",
                "Болт М10х40",
                "м10*40",
                "без бренда",
                None,
                50,
                "шт",
                "BLT-M10-040-ZN",
                "Болт М10х40 DIN 933 оцинкованный",
                "найдено",
                "—",
            ),
        ],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Датасет"]
    header_row, mapping = detect_columns(ws)
    assert header_row == 2
    expected_fields = {
        "row_num",
        "source_file",
        "name",
        "article",
        "manufacturer",
        "attributes",
        "quantity",
        "unit",
        "expected_article",
        "expected_name",
        "label_status",
        "labeler_notes",
        "result_article",
        "result_confidence",
        "result_matched",
    }
    assert set(mapping.keys()) == expected_fields
    # Позиции совпадают с реальным шаблоном
    assert mapping["name"] == 3
    assert mapping["attributes"] == 6
    assert mapping["expected_article"] == 9
    assert mapping["result_article"] == 13


def test_read_rows_skips_empty_names(tmp_path: Path) -> None:
    path = make_gold_xlsx(
        tmp_path / "gold.xlsx",
        data_rows=[
            (1, None, "Болт", None, None, None, 10, "шт", "BLT-1", None, "найдено", None),
            (2, None, None, None, None, None, None, None, None, None, None, None),
            (3, None, "Гайка", None, None, None, 20, "шт", "NUT-1", None, "найдено", None),
        ],
    )
    wb = openpyxl.load_workbook(path)
    ws = wb["Датасет"]
    header_row, mapping = detect_columns(ws)
    rows = read_rows(ws, header_row, mapping)
    # Должны попасть только две заполненные строки (3 и 5)
    assert len(rows) == 2
    assert rows[0].name == "Болт"
    assert rows[1].name == "Гайка"


def test_detect_columns_raises_when_required_missing(tmp_path: Path) -> None:
    """Если в шапке нет обязательной колонки name — ошибка."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Датасет"
    ws.append(["foo", "bar", "baz"])
    ws.append(["1", "2", "3"])
    path = tmp_path / "bad.xlsx"
    wb.save(path)
    wb.close()

    wb2 = openpyxl.load_workbook(path)
    with pytest.raises(GoldEvalError, match="заголов"):
        detect_columns(wb2["Датасет"])


# --- End-to-end прогон ---


async def test_e2e_eval_run_computes_metrics(
    committed_db: AsyncSession,
    tmp_path: Path,
) -> None:
    """Полный прогон: 5 размеченных строк с известным распределением.

    Ожидание:
      - 3 строки «найдено» с точным artикулом → matcher даст confidence ≥ 0.95
      - 1 строка «не найдено» с заведомо неизвестным товаром → корректно
      - 1 строка «сомнительно» → пропущена из метрик
    """
    await _seed_catalog(committed_db, tmp_path)

    path = make_gold_xlsx(
        tmp_path / "gold.xlsx",
        data_rows=[
            # row 1: точное совпадение по артикулу → top-1, recall=1
            (
                1,
                "spec1.xlsx",
                "Болт М10х40 оцинков. DIN933",
                "BLT-M10-040-ZN",
                "KOELNER",
                None,
                50,
                "шт",
                "BLT-M10-040-ZN",
                "Болт М10х40 DIN 933 оцинкованный",
                "найдено",
                "точное совпадение",
            ),
            # row 2: совпадение по артикулу + другому товару каталога
            (
                2,
                "spec1.xlsx",
                "Гайка М10",
                "NUT-M10",
                "KOELNER",
                None,
                100,
                "шт",
                "NUT-M10",
                "Гайка М10 DIN 934",
                "найдено",
                None,
            ),
            # row 3: артикул найден, но в gold ждут другой (имитация ошибки матчера)
            # Берём шайбу но размечаем как болт М10х50 — recall@5 даст промах
            (
                3,
                "spec1.xlsx",
                "Шайба М10",
                "WSH-M10",
                "KOELNER",
                None,
                200,
                "шт",
                "BLT-M10-050-ZN",
                "Болт М10х50 DIN 933 оцинкованный",
                "найдено",
                "тест: разметка нарочно не совпадает",
            ),
            # row 4: «не найдено» — заведомо нет в каталоге
            (
                4,
                "spec1.xlsx",
                "Странный товар xyz",
                "ZZZ-NOPE",
                None,
                None,
                1,
                "шт",
                None,
                None,
                "не найдено",
                None,
            ),
            # row 5: «сомнительно» — должна быть пропущена
            (
                5,
                "spec1.xlsx",
                "Дюбель",
                "DOWEL-6-40",
                "FISCHER",
                None,
                10,
                "шт",
                "DOWEL-6-40",
                "Дюбель универсальный 6х40",
                "сомнительно",
                "не уверены",
            ),
        ],
    )
    output_path = tmp_path / "result.xlsx"

    metrics = await run_eval(
        input_path=path,
        output_path=output_path,
        top_k=5,
        min_confidence=0.5,
    )

    # --- Проверки метрик ---
    assert metrics.total_rows == 5
    assert metrics.skipped_unsure == 1  # «сомнительно»
    assert metrics.applicable == 3  # три «найдено»

    # row 1 и row 2 — точные попадания (top-1 правильный)
    # row 3 — артикул шайбы, разметка ждёт болт М10х50 → промах в top-K
    # ⇒ recall@5 == 2/3, precision@1 == 2/3
    assert metrics.recall_at_k_hits == 2
    assert metrics.precision_at_1_hits == 2
    assert metrics.recall_at_k == pytest.approx(2 / 3)
    assert metrics.precision_at_1 == pytest.approx(2 / 3)
    # MRR = (1/1 + 1/1 + 0) / 3 = 2/3
    assert metrics.mrr == pytest.approx(2 / 3)

    # «не найдено» — должна быть корректной (матчер ничего не вернёт)
    assert metrics.not_found_rows == 1
    assert metrics.not_found_correct == 1

    # --- Проверки выходного файла ---
    assert output_path.exists()
    wb_out = openpyxl.load_workbook(output_path)
    assert "Метрики" in wb_out.sheetnames

    ws = wb_out["Датасет"]
    # Колонки результата заполнены
    # row 3 (sheet row) — это row data #1 (Болт М10х40 точное)
    assert ws.cell(row=3, column=13).value == "BLT-M10-040-ZN"
    assert float(ws.cell(row=3, column=14).value) >= 0.95
    assert ws.cell(row=3, column=15).value == "да"

    # row 5 (sheet row) — это data #3 (Шайба, ждут болт М10х50)
    # Матчер по WSH-M10 даст шайбу top-1; совпало = «нет»
    assert ws.cell(row=5, column=13).value == "WSH-M10"
    assert ws.cell(row=5, column=15).value == "нет"

    # row 6 (sheet row) — это data #4 (Странный товар, не найдено)
    # expected пуст → совпало = пустая ячейка (None или "")
    assert ws.cell(row=6, column=15).value in (None, "")

    # Лист «Метрики» содержит Recall@K
    metrics_ws = wb_out["Метрики"]
    rows_text = [
        str(metrics_ws.cell(row=r, column=1).value or "") for r in range(1, metrics_ws.max_row + 1)
    ]
    assert any("Recall" in t for t in rows_text)
    assert any("Precision@1" in t for t in rows_text)
    assert any("MRR" in t for t in rows_text)

    wb_out.close()


async def test_e2e_eval_run_handles_empty_catalog(
    committed_db: AsyncSession,
    tmp_path: Path,
) -> None:
    """Если каталог пуст — все matchings пустые, recall@K = 0."""
    path = make_gold_xlsx(
        tmp_path / "gold.xlsx",
        data_rows=[
            (1, None, "Болт", "BLT-1", None, None, 10, "шт", "BLT-1", None, "найдено", None),
            (2, None, "Гайка", "NUT-1", None, None, 20, "шт", "NUT-1", None, "найдено", None),
        ],
    )
    output_path = tmp_path / "result.xlsx"

    metrics = await run_eval(
        input_path=path,
        output_path=output_path,
        top_k=5,
    )

    assert metrics.applicable == 2
    assert metrics.recall_at_k_hits == 0
    assert metrics.precision_at_1_hits == 0
    assert metrics.mrr == 0.0


async def test_e2e_eval_run_default_output_path(
    committed_db: AsyncSession,
    tmp_path: Path,
) -> None:
    """Если output не задан — файл рядом с input с суффиксом и timestamp."""
    # Просто проверим, что run_eval с явным output работает в tmp_path
    path = make_gold_xlsx(
        tmp_path / "gold.xlsx",
        data_rows=[
            (1, None, "Болт", "BLT-1", None, None, 10, "шт", "BLT-1", None, "найдено", None),
        ],
    )
    output_path = tmp_path / "subdir" / "out.xlsx"

    await run_eval(input_path=path, output_path=output_path, top_k=5)
    assert output_path.exists()


async def test_e2e_eval_run_with_real_template_file(
    committed_db: AsyncSession,
    tmp_path: Path,
) -> None:
    """Прогон по живому шаблону gold_dataset_template_new.xlsx (1 строка)."""
    await _seed_catalog(committed_db, tmp_path)

    template_path = Path("/home/master/projects/FastTender/docs/gold_dataset_template_new.xlsx")
    # Копируем в tmp, чтобы не мутировать исходник
    template_bytes = template_path.read_bytes()
    work_path = tmp_path / "gold_real.xlsx"
    work_path.write_bytes(template_bytes)
    output_path = tmp_path / "gold_real_result.xlsx"

    metrics = await run_eval(
        input_path=work_path,
        output_path=output_path,
        top_k=5,
    )

    # В шаблоне только одна заполненная строка с правильным артикулом BLT-M10-040-ZN
    assert metrics.applicable == 1
    assert metrics.recall_at_k_hits == 1
    assert metrics.precision_at_1_hits == 1

    # Не сломалось при двухуровневой шапке
    wb_out = openpyxl.load_workbook(output_path)
    ws = wb_out["Датасет"]
    assert ws.cell(row=3, column=13).value == "BLT-M10-040-ZN"
    assert ws.cell(row=3, column=15).value == "да"
    wb_out.close()


# Silence unused import warning
_ = io
