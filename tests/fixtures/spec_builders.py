"""Генераторы тестовых файлов спецификаций — XLSX и CSV.

Файлы создаются в tmp_path во время теста, в репозитории не хранятся —
это и компактнее, и нагляднее (структура файла видна прямо в коде теста).
"""

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


def make_xlsx(
    path: Path,
    rows: list[list[Any]],
    *,
    sheet_name: str = "Спецификация",
    merged_ranges: list[str] | None = None,
) -> Path:
    """Создаёт XLSX-файл с указанными строками и объединёнными диапазонами.

    Args:
        path: куда сохранить.
        rows: содержимое строк (как list[list]).
        sheet_name: имя листа.
        merged_ranges: список a1-нотаций для объединения, напр. ["A1:D1"].

    Returns:
        path для удобства цепочки.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    if merged_ranges:
        for rng in merged_ranges:
            ws.merge_cells(rng)
            # Проверяем валидность диапазона (для будущего, если будем менять)
            range_boundaries(rng)

    # Подгоняем ширину колонок — это не на парсер, просто чтоб файл был «реальный»
    for c_idx in range(1, (max(len(r) for r in rows) if rows else 1) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 18

    wb.save(path)
    wb.close()
    return path


def make_csv(
    path: Path,
    rows: list[list[Any]],
    *,
    encoding: str = "utf-8",
    delimiter: str = ",",
) -> Path:
    """Создаёт CSV-файл в заданной кодировке и с заданным разделителем."""
    import csv

    with path.open("w", encoding=encoding, newline="") as f:
        writer = csv.writer(f, delimiter=delimiter)
        for row in rows:
            writer.writerow(["" if v is None else str(v) for v in row])
    return path
